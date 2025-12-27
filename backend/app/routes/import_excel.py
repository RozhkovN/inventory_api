# app/routes/import_excel.py
import io
import re
from decimal import Decimal
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from app.database import get_db
from app.models.product import Product

router = APIRouter()

def parse_price_and_coeff(cell_value):
    if isinstance(cell_value, str) and cell_value.startswith('='):
        match = re.search(r'=\s*(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)', cell_value)
        if match:
            try:
                base = Decimal(match.group(1))
                coeff = Decimal(match.group(2))
                return base, coeff
            except:
                pass
    try:
        price = Decimal(str(cell_value).replace(',', '.'))
        return price, Decimal('1.0')
    except:
        return Decimal('0'), Decimal('1.0')

@router.post("/import/warehouse", summary="Импорт склада из Excel (лист 'Склад')")
async def upload_warehouse_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Только .xlsx файлы")

    contents = await file.read()

    try:
        wb = load_workbook(io.BytesIO(contents), data_only=False)
        ws = wb["Склад"]
    except KeyError:
        raise HTTPException(400, "В файле нет листа 'Склад'")
    except Exception as e:
        raise HTTPException(400, f"Ошибка чтения Excel: {e}")

    created = 0
    # Читаем 5 колонок: A, B, C, D, E
    for row in ws.iter_rows(min_row=2, max_col=5):
        try:
            name_cell = row[1]  # B
            qty_cell = row[2]   # C
            price_cell = row[3] # D
            sku_cell = row[4]   # E

            if not name_cell.value or not qty_cell.value:
                continue

            name = str(name_cell.value).strip()
            sku = str(sku_cell.value).strip() if sku_cell.value else None
            qty = int(qty_cell.value)
            purchase_price, coefficient = parse_price_and_coeff(price_cell.value)

            # Ищем товар по (name, sku) — именно так они уникальны!
            existing = db.query(Product).filter(
                Product.name == name,
                Product.sku == sku
            ).first()

            if existing:
                existing.quantity += qty
                db.add(existing)
            else:
                product = Product(
                    name=name,
                    sku=sku,
                    purchase_price=purchase_price,
                    coefficient=coefficient,
                    quantity=qty
                )
                db.add(product)
                created += 1

        except Exception as e:
            print(f"Ошибка строки: {e}")
            continue

    db.commit()
    return {"message": "Склад успешно импортирован", "products_created": created}