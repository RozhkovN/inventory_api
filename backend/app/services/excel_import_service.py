# app/services/excel_import_service.py
import io
import re
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.models.product import Product

def parse_price_and_coeff(cell_value, raw_value):
    if isinstance(raw_value, str) and raw_value.startswith('='):
        match = re.search(r'=\s*(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)', raw_value)
        if match:
            return float(match.group(1)), float(match.group(2))
    return float(cell_value), 1.0

def import_warehouse_from_excel(db: Session, file_bytes: bytes):
    try:
        wb_raw = load_workbook(io.BytesIO(file_bytes), data_only=False)
        wb_vals = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws_raw = wb_raw["Склад"]
        ws_vals = wb_vals["Склад"]
    except KeyError:
        raise ValueError("В файле нет листа 'Склад'")
    except Exception as e:
        raise ValueError(f"Ошибка чтения Excel: {e}")

    created = 0
    for (row_raw, row_vals) in zip(ws_raw.iter_rows(min_row=2), ws_vals.iter_rows(min_row=2)):
        try:
            name_raw = row_raw[1]
            qty_raw = row_vals[2]
            price_val = row_vals[3]
            price_raw = row_raw[3]

            if not name_raw.value or qty_raw.value is None or price_val.value is None:
                continue

            name = str(name_raw.value).strip()
            qty = int(qty_raw.value)
            purchase_price, coefficient = parse_price_and_coeff(price_val.value, price_raw.value)

            existing = db.query(Product).filter(Product.name == name).first()
            if existing:
                existing.quantity += qty
                db.add(existing)
            else:
                product = Product(
                    name=name,
                    purchase_price=purchase_price,
                    coefficient=coefficient,
                    quantity=qty
                )
                db.add(product)
                created += 1
        except Exception as e:
            print(f"Ошибка при импорте строки: {e}")
            continue

    db.commit()
    return {"products_created": created}