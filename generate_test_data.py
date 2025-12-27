# generate_test_data.py
import openpyxl
from openpyxl.styles import Font
import random

# Базовые названия
base_names = [
    "ТОНЕР-КАРТРИДЖ ЧЕРНЫЙ",
    "ТОНЕР-КАРТРИДЖ СИНИЙ",
    "ТОНЕР-КАРТРИДЖ ПУРПУРНЫЙ",
    "ТОНЕР-КАРТРИДЖ ЖЕЛТЫЙ",
    "ДРАМ-КАРТРИДЖ ЧЕРНЫЙ",
    "ФИЛЬТР ОЗОНОВЫЙ",
    "НАГРЕВАТЕЛЬНЫЙ ВАЛ",
    "ПРИЖИМНОЙ ВАЛ",
    "РЕМЕНЬ ПЕРЕНОСА",
    "ЧИП НА ДРАМ"
]

models = ["700DCP", "C1000", "V80", "C60", "DC250", "C75", "WC4110", "C560", "D110", "PrimeLink C9070"]
colors = ["BLK", "CYN", "MAG", "YEL"]
types = ["TN", "DM", "CH", "FLT", "ROL"]  # Типы для SKU

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Склад"

# Заголовки
headers = ["#", "Клиент/товар", "количество", "закупка за штуку", "SKU"]
for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header).font = Font(bold=True)

used_skus = set()

for i in range(1, 101):
    base = random.choice(base_names)
    model = random.choice(models)
    name = f"{base} {model}"
    
    qty = random.randint(1, 20)
    purchase = round(random.uniform(1000, 30000), 2)
    coeff = round(random.choice([1.0, 1.1, 1.2, 1.3, 1.5]), 1)
    formula = f"={purchase}*{coeff}"

    # Генерируем уникальный SKU, основанный на названии и модели
    base_clean = base.replace("ТОНЕР-КАРТРИДЖ ", "").replace("ДРАМ-КАРТРИДЖ ", "")
    color_code = ""
    if "ЧЕРН" in base_clean:
        color_code = "BLK"
    elif "СИНИЙ" in base_clean:
        color_code = "CYN"
    elif "ПУРПУРН" in base_clean:
        color_code = "MAG"
    elif "ЖЕЛТ" in base_clean:
        color_code = "YEL"
    else:
        color_code = "MISC"

    model_clean = model.replace(" ", "").replace("/", "").replace("PrimeLink", "PL")
    base_type = "TN" if "ТОНЕР" in base else ("DM" if "ДРАМ" in base else "OTH")
    sku_base = f"{model_clean}-{color_code}-{base_type}"
    
    # Делаем уникальным (на случай дублей)
    sku = sku_base
    counter = 1
    while sku in used_skus:
        sku = f"{sku_base}-{counter}"
        counter += 1
    used_skus.add(sku)

    # Записываем строку
    ws.cell(row=i+1, column=1, value=i)
    ws.cell(row=i+1, column=2, value=name)
    ws.cell(row=i+1, column=3, value=qty)
    ws.cell(row=i+1, column=4, value=formula)
    ws.cell(row=i+1, column=5, value=sku)

wb.save("test_warehouse.xlsx")
print("✅ test_warehouse.xlsx создан (100 позиций с уникальными SKU)")